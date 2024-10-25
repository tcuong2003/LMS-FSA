from django.shortcuts import render, get_object_or_404, redirect
from .models import Course, Enrollment, ReadingMaterial, Completion, Session, SessionCompletion, Topic, Tag, CourseMaterial
from .forms import CourseForm, EnrollmentForm, CourseSearchForm, SessionForm, TopicForm, TagForm, ReadingMaterialEditForm
from module_group.models import ModuleGroup
from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.contrib import messages
import os
from django.http import FileResponse,Http404
from django.utils.text import slugify
from django.urls import reverse
from feedback.models import CourseFeedback
from .forms import ExcelImportForm
from django.http import HttpResponse
import openpyxl
import pandas as pd
from user.models import User
from django.views.decorators.http import require_POST
from django.http import JsonResponse
from django.core.paginator import Paginator
from datetime import datetime
import base64
from itertools import zip_longest
import numpy as np
import fitz
from django.core.files.storage import default_storage
import random


def export_course(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=lms_course.xlsx'

    workbook = openpyxl.Workbook()

    # Course sheet
    course_worksheet = workbook.active
    course_worksheet.title = 'Course'
    course_columns = ['course_name', 'course_code', 'description', 'creator', 'instructor', 'published', 'prerequisites']
    course_worksheet.append(course_columns)

    for course in Course.objects.all():
        prerequisites_list = ', '.join([prerequisite.course_name for prerequisite in course.prerequisites.all()]) or None
        course_worksheet.append([
            course.course_name,
            course.course_code,
            course.description,
            course.creator.username if course.creator else None,
            course.instructor.username if course.instructor else None,
            course.published,
            prerequisites_list
        ])

    # Session sheet
    session_worksheet = workbook.create_sheet(title='Session')
    session_columns = ['id', 'course_name', 'session_name', 'session_order']
    session_worksheet.append(session_columns)

    for session in Session.objects.all():
        session_worksheet.append([
            session.id,  # Include session ID
            session.course.course_name if session.course else None,
            session.name,
            session.order
        ])

    # Reading Material sheet
    material_worksheet = workbook.create_sheet(title='Reading Material')
    material_columns = ['session_id', 'title', 'material_type', 'order', 'content']
    material_worksheet.append(material_columns)

    for material in ReadingMaterial.objects.all():
        material_worksheet.append([
            material.session.id if material.session else None,
            material.title,
            material.material_type,
            material.order,
            material.content  # Depending on how you want to handle HTML content, you might want to sanitize it or limit the length.
        ])

    # Save the workbook to the response
    workbook.save(response)
    return response


def to_none_if_nan(value):
    """Convert value to None if it is NaN."""
    return None if isinstance(value, float) and np.isnan(value) else value

def import_courses(request):
    if request.method == 'POST':
        form = ExcelImportForm(request.POST, request.FILES)
        if form.is_valid():
            uploaded_file = request.FILES['excel_file']
            try:
                # Read necessary sheets
                course_df = pd.read_excel(uploaded_file, sheet_name='Course')
                session_df = pd.read_excel(uploaded_file, sheet_name='Session')
                material_df = pd.read_excel(uploaded_file, sheet_name='Reading Material')

                # Import courses
                course_imported = 0
                course_updated = 0

                for index, row in course_df.iterrows():
                    course_name = row['course_name']
                    course_code = row['course_code']
                    description = row['description']
                    creator_username = to_none_if_nan(row.get('creator'))
                    instructor_username = to_none_if_nan(row.get('instructor'))
                    prerequisites = to_none_if_nan(row.get('prerequisites'))

                    # Fetch User instances
                    creator = User.objects.filter(username=creator_username).first() if creator_username else None
                    instructor = User.objects.filter(username=instructor_username).first() if instructor_username else None

                    # Get or create the course
                    course, created = Course.objects.get_or_create(
                        course_name=course_name,
                        defaults={
                            'course_code': course_code,
                            'description': description,
                            'creator': creator,
                            'instructor': instructor,
                        }
                    )

                    if created:
                        course_imported += 1
                    else:
                        course_updated += 1

                    # Handle prerequisites
                    if prerequisites:
                        prerequisite_names = [prerequisite.strip() for prerequisite in prerequisites.split(',')]
                        course.prerequisites.clear()
                        for prerequisite_name in prerequisite_names:
                            prerequisite = Course.objects.filter(course_name=prerequisite_name).first()
                            if prerequisite:
                                course.prerequisites.add(prerequisite)
                            else:
                                messages.warning(request, f"Prerequisite '{prerequisite_name}' does not exist for course '{course_name}'.")

                # Import sessions and create mapping
                session_mapping = {}
                for index, row in session_df.iterrows():
                    course_name = row['course_name']
                    session_name = row['session_name']
                    session_order = row['session_order']

                    course = Course.objects.filter(course_name=course_name).first()
                    if course:
                        session, _ = Session.objects.get_or_create(
                            course=course,
                            name=session_name,
                            defaults={'order': session_order}
                        )
                        session_mapping[row['id']] = session  # Map original ID to the new session instance

                # Import reading materials
                material_imported = 0
                for index, row in material_df.iterrows():
                    original_session_id = row['session_id']  # Use the original session ID from the Excel file
                    title = row['title']
                    order = row['order']
                    content = row['content']  # Assume this is HTML content

                    session = session_mapping.get(original_session_id)  # Fetch the session using the mapping
                    if session:
                        ReadingMaterial.objects.get_or_create(
                            session=session,
                            title=title,
                            defaults={
                                'content': content,
                                'order': order,
                            }
                        )
                        material_imported += 1
                    else:
                        messages.warning(request, f"Session ID '{original_session_id}' does not exist for material '{title}'.")

                messages.success(request, f"{course_imported} courses imported successfully! {course_updated} courses already existed. {material_imported} reading materials imported.")
            except Exception as e:
                messages.error(request, f"An error occurred during import: {e}")

            return redirect('course:course_list')
    else:
        form = ExcelImportForm()

    return render(request, 'course/course_list.html', {'form': form})


@login_required
def course_enroll(request, pk):
    course = get_object_or_404(Course, pk=pk)

    # Automatically enroll and redirect to the course detail page
    form = EnrollmentForm(request.POST)

    if form.is_valid():
        enrollment = form.save(commit=False)

        # Fetch prerequisite courses from the Course model
        prerequisite_courses = course.prerequisites.all()

        # Check if the user is enrolled in all prerequisite courses
        enrolled_courses = Enrollment.objects.filter(
            student=request.user,
            course__in=prerequisite_courses
        ).values_list('course', flat=True)

        if all(prereq.id in enrolled_courses for prereq in prerequisite_courses):
            enrollment.student = request.user
            enrollment.course = course
            enrollment.save()
            messages.success(request, f'You have been enrolled in {course.course_name}.')
        else:
            messages.error(request, 'You do not meet the prerequisites for this course.')
            return redirect('course:course_list')  # Redirect to course list or another page

    return redirect('course:course_detail', pk=course.pk)


@login_required
def course_unenroll(request, pk):
    course = get_object_or_404(Course, pk=pk)
    enrollment = Enrollment.objects.filter(student=request.user, course=course).first()

    if request.method == 'POST':
        # Unenroll the user and redirect to course list with a message
        if enrollment:
            enrollment.delete()
            messages.success(request, f'You have been unenrolled from {course.course_name}.')
        return redirect('course:course_list')

    # Render confirmation page
    return render(request, 'course/course_unenroll.html', {'course': course})

def course_list(request):
    if request.user.is_superuser:
        # Superuser can see all courses
        courses = Course.objects.all()
    elif Course.objects.filter(instructor=request.user).exists():
        # Instructors can see all courses they are teaching, published or not
        courses = Course.objects.filter(
            Q(published=True) | Q(instructor=request.user)
        )
    else:
        courses = Course.objects.filter(published=True)  # Other users see only published courses

    module_groups = ModuleGroup.objects.all()
    enrollments = Enrollment.objects.filter(student=request.user)
    enrolled_courses = {enrollment.course.id for enrollment in enrollments}
    is_instructor = Course.objects.filter(instructor=request.user).exists()

    # Calculate completion percentage for each course
    for course in courses:
        course.completion_percent = course.get_completion_percent(request.user)

    # Recommended courses logic
    recommended_courses = []
    for course in courses:
        if course.id not in enrolled_courses:
            for enrolled_course_id in enrolled_courses:
                enrolled_course = Course.objects.get(id=enrolled_course_id)
                enrolled_tags = set(enrolled_course.tags.all()) if enrolled_course.tags else set()
                current_tags = set(course.tags.all()) if course.tags else set()

                # Calculate the similarity
                if enrolled_tags:
                    shared_tags = enrolled_tags.intersection(current_tags)
                    similarity = len(shared_tags) / len(enrolled_tags)
                    if similarity >= 0.6:  # 60% similarity
                        recommended_courses.append(course)
                        break  # No need to check other enrolled courses

    # 14/10/2024 - New addition for recommended courses pagination
    recommended_paginator = Paginator(recommended_courses, 5)  # Show 5 recommended courses per page
    recommended_page_number = request.GET.get('recommended_page')
    recommended_page_obj = recommended_paginator.get_page(recommended_page_number)

    # 14/10/2024
    current_recommended_page = recommended_page_obj.number if recommended_page_obj else 1

    # Pagination for main courses
    paginator = Paginator(courses, 3)  # Show 10 courses per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'course/course_list.html', {
        'module_groups': module_groups,
        'page_obj': page_obj,  # Pagination object for template
        'courses': page_obj,  # Consistent with template expectations
        'enrolled_courses': enrolled_courses,  # To show enrolled status
        'recommended_courses': recommended_page_obj,
        'current_recommended_page': current_recommended_page,  # Pass the current page number
        'is_instructor': is_instructor,
    })


def course_add(request):
    if request.method == 'POST':
        course_form = CourseForm(request.POST, request.FILES)

        if course_form.is_valid():
            # Save the course
            course = course_form.save(commit=False)
            course.creator = request.user

            # Handle image upload
            if 'image' in request.FILES:
                course.image = request.FILES['image']

            course.save()

            # Handle prerequisite courses
            prerequisite_ids = request.POST.getlist('prerequisite_courses[]')
            for prerequisite_id in prerequisite_ids:
                if prerequisite_id:
                    prerequisite_course = Course.objects.get(id=prerequisite_id)
                    course.prerequisites.add(prerequisite_course)

            # Create sessions for the course directly
            session_name = request.POST.get('session_name')
            session_quantity = int(request.POST.get('session_quantity', 0))
            if session_name and session_quantity > 0:
                for i in range(1, session_quantity + 1):
                    session = Session(
                        course=course,
                        name=f"{session_name}{i}",
                        order=i
                    )
                    session.save()

            messages.success(request, 'Course and sessions created successfully.')
            return redirect('course:course_list')
        else:
            messages.error(request, 'There was an error creating the course. Please check the form.')

    else:
        course_form = CourseForm()
        session_form = SessionForm()

    all_courses = Course.objects.all()

    return render(request, 'course/course_form.html', {
        'course_form': course_form,
        'session_form': session_form,
        'all_courses': all_courses,
    })

# course/views.py
def course_edit(request, pk):
    course = get_object_or_404(Course, pk=pk)
    all_courses = Course.objects.exclude(id=course.id)

    if request.method == 'POST':
        course_form = CourseForm(request.POST, request.FILES, instance=course)

        if course_form.is_valid():
            course = course_form.save(commit=False)
            course.creator = request.user

            # Debug: Check if an image has been uploaded
            if 'image' in request.FILES:
                print("Image uploaded:", request.FILES['image'])

                # Only delete old image after saving the new one
                new_image = request.FILES['image']
                if course.image.name != new_image.name and course.image.path:
                    print("Old image exists, deleting:", course.image.path)
                    default_storage.delete(course.image.path)
                    course.image.delete()
                course.image = new_image  # Assign the new image to the course

                # Save course first with the new image to avoid premature file closing
                course.save()

            elif request.POST.get('delete_image') == 'on':  # Explicitly check for 'on'
                print("Delete image checkbox is checked")

                if course.image:
                    print("Deleting old image:", course.image.path)
                    default_storage.delete(course.image.path)
                    course.image.delete()

            course.save()

            # Handle tag deletion
            current_tags = list(course.tags.all())
            for tag in current_tags:
                if request.POST.get(f'delete_tag_{tag.id}'):
                    print(f"Deleting tag: {tag.name}")
                    course.tags.remove(tag)

            # Handle adding new tags
            tag_ids = request.POST.getlist('tags')
            print("Retrieved tag_ids from POST:", tag_ids)
            for tag_id in tag_ids:
                if tag_id:
                    print("Processing tag_id:", tag_id)
                    tag = Tag.objects.get(id=tag_id)
                    course.tags.add(tag)

            # Handle prerequisite deletion
            current_prerequisites = request.POST.get('deleted_prerequisite_ids')
            if current_prerequisites:
                current_prerequisites = current_prerequisites.split(',')
                for prereq_id in current_prerequisites:
                    if prereq_id:  # Ensure there's a value to work with
                        try:
                            # Convert the ID to an integer
                            prereq_id = int(prereq_id)
                            # Attempt to get the prerequisite from the course
                            prereq = course.prerequisites.get(id=prereq_id)
                            course.prerequisites.remove(prereq)
                            print(f"Deleted prerequisite: {prereq.course_name}")
                        except course.prerequisites.DoesNotExist:
                            print(f"Prerequisite with ID {prereq_id} does not exist.")

            # Handle adding new prerequisites
            prerequisite_ids = request.POST.getlist('prerequisite_courses')
            for prerequisite_id in prerequisite_ids:
                if prerequisite_id:
                    prerequisite_course = Course.objects.get(id=prerequisite_id)
                    print(f"Adding prerequisite: {prerequisite_course.course_name}")
                    course.prerequisites.add(prerequisite_course)

            # Handle existing sessions update
            session_ids = request.POST.getlist('session_ids')
            session_names = request.POST.getlist('session_names')
            for session_id, session_name in zip(session_ids, session_names):
                if session_id:  # Only process if session_id is not empty
                    print(f"Updating session {session_id} name to: {session_name}")
                    session = Session.objects.get(id=session_id)
                    session.name = session_name
                    session.save()

            # Handle adding new sessions
            new_session_names = request.POST.getlist('new_session_names')
            for session_name in new_session_names:
                if session_name:
                    print(f"Adding new session: {session_name}")
                    Session.objects.create(course=course, name=session_name, order=course.sessions.count() + 1)

            # Handle session deletion
            delete_session_ids = request.POST.get('delete_session_ids')
            if delete_session_ids:
                delete_session_ids = delete_session_ids.split(',')
                for session_id in delete_session_ids:
                    if session_id:
                        Session.objects.filter(id=session_id).delete()

            messages.success(request, 'course updated successfully.')
            return redirect('course:course_edit', pk=course.pk)
        else:
            print("Form is not valid")
            print(course_form.errors)
            messages.error(request, 'There was an error updating the course. Please check the form.')

    else:
        course_form = CourseForm(instance=course)
        prerequisites = course.prerequisites.all()
        sessions = course.sessions.all()
        topics = Topic.objects.all()
        tags = Tag.objects.all()

    return render(request, 'course/edit_form.html', {
        'course_form': course_form,
        'course': course,
        'prerequisites': prerequisites,
        'all_courses': all_courses,
        'sessions': sessions,  # Pass sessions to template
        'topics': topics,
        'tags': tags,
    })

def course_delete(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.method == 'POST':
        course.delete()
        return redirect('course:course_list')
    return render(request, 'course/course_confirm_delete.html', {'course': course})

@login_required
def course_detail(request, pk):
    # Get the course based on the primary key (pk)
    course = get_object_or_404(Course, pk=pk)

    # Get related documents and videos
    is_enrolled = Enrollment.objects.filter(student=request.user, course=course).exists()
    users_enrolled_count = Enrollment.objects.filter(course=course).count()

    # Get all feedback related to the course
    feedbacks = CourseFeedback.objects.filter(course=course)

    # Calculate the course's average rating
    if feedbacks.exists():
        total_rating = sum(feedback.average_rating() for feedback in feedbacks)
        course_average_rating = total_rating / feedbacks.count()
    else:
        course_average_rating = None  # No feedback yet

    if course_average_rating is not None:
        course_average_rating_star = course_average_rating * 100 / 5
    else:
        course_average_rating_star = 0

        # Get prerequisite courses directly from the course's `prerequisites` column
    prerequisites = course.prerequisites.all()

    sessions = Session.objects.filter(course=course)
    session_count = sessions.count()
    # Get 5 random tags
    all_tags = list(course.tags.all())
    random_tags = random.sample(all_tags, min(4, len(all_tags)))

    # Fetch the 5 newest feedback entries for this course
    latest_feedbacks = CourseFeedback.objects.filter(course=course).order_by('-created_at')[:5]


    # Get all users who are instructors (you might need to adjust this query based on how you identify instructors)
    instructor = course.instructor  # Assuming instructors are staff members
    is_instructor = Course.objects.filter(instructor=request.user).exists()
    if is_instructor:
        user_type = 'instructor'
    else:
        user_type = 'student'

    enrolled_users = Enrollment.objects.filter(course=course).select_related('student')

    # Calculate progress for each enrolled user
    user_progress = [
        {
            'user': enrollment.student,
            'progress': course.get_completion_percent(enrollment.student)
        }
        for enrollment in enrolled_users
    ]

    context = {
        'course': course,
        'prerequisites': prerequisites,
        'is_enrolled': is_enrolled,
        'users_enrolled_count': users_enrolled_count,
        'course_average_rating_star': course_average_rating_star,
        'course_average_rating': course_average_rating,
        'feedbacks': feedbacks,
        'sessions': sessions,
        'session_count': session_count,
        'latest_feedbacks': latest_feedbacks,
        'tags': course.tags.all() if course.tags else [],
        'instructor': instructor,  # Add this line
        'user_type': user_type,
        'user_progress': user_progress,
        'random_tags': random_tags,
    }

    return render(request, 'course/course_detail.html', context)

def users_enrolled(request, pk):
    # Lấy môn học dựa trên khóa chính (primary key)
    course = get_object_or_404(Course, pk=pk)

    # Lấy danh sách người dùng đã đăng ký môn học
    enrolled_users = Enrollment.objects.filter(course=course).select_related('student')

    # Calculate progress for each enrolled user
    user_progress = [
        {
            'user': enrollment.student,
            'progress': course.get_completion_percent(enrollment.student)
        }
        for enrollment in enrolled_users
    ]

    return render(request, 'course/users_course_enrolled.html', {
        'course': course,
        'user_progress': user_progress,
        'enrolled_users': enrolled_users,
    })

def course_search(request):
    form = CourseSearchForm(request.GET or None)
    query = request.GET.get('query', '')
    courses = Course.objects.all()

    if query:
        courses = courses.filter(
            Q(course_name__icontains=query) |
            Q(description__icontains=query) |
            Q(course_code__icontains=query))

    # Add pagination for search results
    paginator = Paginator(courses, 10)  # Show 10 results per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'form': form,
        'page_obj': page_obj,  # For paginated results
        'courses': page_obj,  # Pass the paginated courses as 'courses' for template consistency
    }
    return render(request, 'course/course_list.html', context)

@login_required
def reorder_course_materials(request, pk, session_id):
    # Fetch the course
    course = get_object_or_404(Course, pk=pk)

    # Fetch all sessions related to the course
    sessions = Session.objects.filter(course=course)

    # Fetch materials for the selected session, defaulting to the first session
    selected_session_id = request.POST.get('session_id') or session_id
    session = get_object_or_404(Session, id=selected_session_id)
    materials = CourseMaterial.objects.filter(session=session).order_by('order')

    if request.method == 'POST':
        # Check if the request is for reordering materials
        if 'order' in request.POST:
            for material in materials:
                new_order = request.POST.get(f'order_{material.id}')
                if new_order:
                    material.order = int(new_order)  # Convert to integer
                    material.save()

            success_message = "Order updated successfully!"
            return render(request, 'material/reorder_course_material.html', {
                'course': course,
                'sessions': sessions,
                'materials': materials,
                'selected_session_id': selected_session_id,
                'success_message': success_message,
            })

    # Pass the course, sessions, and materials to the template
    return render(request, 'material/reorder_course_material.html', {
        'course': course,
        'sessions': sessions,
        'materials': materials,
        'selected_session_id': selected_session_id,
    })
def reading_material_detail(request, id):
    # Fetch the reading material by ID or return a 404 if it doesn't exist
    reading_material = get_object_or_404(ReadingMaterial, id=id)
    return render(request, 'material/reading_material_detail.html', {'reading_material': reading_material})


def edit_reading_material(request, pk, session_id, reading_material_id):
    # Retrieve the course
    course = get_object_or_404(Course, pk=pk)
    # Retrieve all sessions for the course
    sessions = Session.objects.filter(course=course)

    # Get the selected session
    selected_session_id = session_id  # Use session_id parameter
    session = get_object_or_404(Session, id=selected_session_id)

    # Retrieve the reading material to edit
    reading_material = get_object_or_404(ReadingMaterial, id=reading_material_id)

    if request.method == 'POST':
        form = ReadingMaterialEditForm(request.POST, instance=reading_material)
        if form.is_valid():
            form.save()
            messages.success(request, 'Reading material updated successfully.')
            return redirect('course:course_content_edit', pk=pk, session_id=session_id)
    else:
        form = ReadingMaterialEditForm(instance=reading_material)

    context = {
        'reading_material': reading_material,
        'form': form,
        'course': course,  # Pass the course to the context
        'sessions': sessions,  # Pass the sessions to the context (if needed)
        'session': session,  # Pass the selected session to the context
    }

    return render(request, 'material/edit_reading_material.html', context)

@login_required
def course_content(request, pk, session_id):
    course = get_object_or_404(Course, pk=pk)
    sessions = Session.objects.filter(course=course).order_by('order')

    selected_session_id = request.POST.get('session_id') or session_id

    current_session = get_object_or_404(Session, id=selected_session_id)

    materials = CourseMaterial.objects.filter(session=current_session).order_by('order')

    file_id = request.GET.get('file_id')
    file_type = request.GET.get('file_type')
    current_material = None
    if file_id and file_type:
        try:
            current_material = CourseMaterial.objects.get(id=file_id, material_type=file_type, session=current_session)
        except CourseMaterial.DoesNotExist:
            current_material = materials.first() if materials.exists() else None
    else:
        current_material = materials.first() if materials.exists() else None

    next_material = materials.filter(order__gt=current_material.order).first() if current_material else None
    next_session = None

    if not next_material:
        next_session = Session.objects.filter(course=course, order__gt=current_session.order).order_by('order').first()
        if next_session:
            next_material = CourseMaterial.objects.filter(session=next_session).order_by('order').first()

    content_type = None
    preview_content = None

    if current_material:
        if current_material.material_type == 'assignments':
            reading = ReadingMaterial.objects.get(material_id=current_material.id)
            preview_content = reading.content
            content_type = 'assignments'
        elif current_material.material_type == 'labs':
            reading = ReadingMaterial.objects.get(material_id=current_material.id)
            preview_content = reading.content
            content_type = 'labs'
        elif current_material.material_type == 'lectures':
            reading = ReadingMaterial.objects.get(material_id=current_material.id)
            preview_content = reading.content
            content_type = 'lectures'
        elif current_material.material_type == 'references':
            reading = ReadingMaterial.objects.get(material_id=current_material.id)
            preview_content = reading.content
            content_type = 'references'

    completion_status = Completion.objects.filter(
        session=current_session,
        material=current_material,
        user=request.user,
        completed=True
    ).exists() if current_material else False

    total_materials = CourseMaterial.objects.filter(session__course=course).count()
    completed_materials = Completion.objects.filter(
        session__course=course,
        user=request.user,
        completed=True
    ).count()
    completion_percent = (completed_materials / total_materials) * 100 if total_materials > 0 else 0

    total_sessions = sessions.count()
    completed_sessions = SessionCompletion.objects.filter(course=course, user=request.user, completed=True).count()

    certificate_url = None
    if total_sessions > 0 and completed_sessions == total_sessions:
        # Call the function to generate the certificate URL
        certificate_url = reverse('course:generate_certificate', kwargs={'pk': course.pk})

    context = {
        'course': course,
        'sessions': sessions,
        'current_session': current_session,
        'materials': materials,
        'current_material': current_material,
        'next_material': next_material,
        'content_type': content_type,
        'preview_content': preview_content,
        'completion_status': completion_status,
        'completion_percent': completion_percent,
        'certificate_url': certificate_url,
        'next_session': next_session,
    }

    return render(request, 'course/course_content.html', context)


@require_POST
@login_required
def toggle_completion(request, pk):
    course = get_object_or_404(Course, pk=pk)
    file_id = request.POST.get('file_id')

    material = get_object_or_404(CourseMaterial, id=file_id, session__course=course)
    session = material.session

    completion, created = Completion.objects.get_or_create(
        session=session,
        material=material,
        user=request.user,
    )
    completion.completed = not completion.completed
    completion.save()

    # Check if all materials in the session are completed
    total_materials = session.materials.count()
    completed_materials = Completion.objects.filter(session=session, user=request.user, completed=True).count()
    session_completed = total_materials == completed_materials

    SessionCompletion.objects.update_or_create(
        user=request.user,
        session=session,
        course=course,
        defaults={'completed': session_completed}
    )

    # Find the next item
    next_material = CourseMaterial.objects.filter(
        session=session,
        order__gt=material.order
    ).order_by('order').first()

    next_session = None
    if not next_material:
        next_session = Session.objects.filter(course=course, order__gt=session.order).order_by('order').first()
        if next_session:
            next_material = CourseMaterial.objects.filter(session=next_session).order_by('order').first()

    next_item_type = next_material.material_type if next_material else None
    next_item_id = next_material.id if next_material else None
    next_session_id = next_session.id if next_session else None

    return JsonResponse({
        'completed': completion.completed,
        'next_item_type': next_item_type,
        'next_item_id': next_item_id,
        'next_session_id': next_session_id
    })
# In course/views.py

@login_required
def course_content_edit(request, pk, session_id):
    course = get_object_or_404(Course, pk=pk)
    sessions = Session.objects.filter(course=course)

    # Default to the first session if not specified in POST
    selected_session_id = request.POST.get('session_id') or session_id
    session = get_object_or_404(Session, id=selected_session_id)

    # Fetch materials associated with the selected session
    materials = CourseMaterial.objects.filter(session=session)
    reading_materials = ReadingMaterial.objects.filter(material__in=materials)

    if request.method == 'POST':
        # Process reading materials for deletion
        for reading_material in reading_materials:
            if f'delete_reading_material_{reading_material.id}' in request.POST:
                reading_material.delete()

        # Handle uploaded PDF
        if 'uploaded_material_file[]' in request.FILES and 'uploaded_material_type[]' in request.POST:
            uploaded_files = request.FILES.getlist('uploaded_material_file[]')
            one_material_type = request.POST.get('uploaded_material_type[]')  # Get the selected material type (single value)
            material_types = [one_material_type] * len(uploaded_files)

            for uploaded_file, material_type in zip(uploaded_files, material_types):
                file_name = os.path.splitext(uploaded_file.name)[0]
            # Read the PDF and extract text
                extracted_content = ""
                with fitz.open(stream=uploaded_file.read(), filetype='pdf') as pdf:
                    for page in pdf:
                        # Extract text and formatting
                        extracted_content += page.get_text("html")  # Extract as HTML

                if extracted_content.strip():
                    reading_material = ReadingMaterial.objects.create(
                        title=file_name,  # Use the uploaded file name as the title
                        content=extracted_content,
                    )
                    course_material = CourseMaterial.objects.create(
                        session=session,
                        material_id=reading_material.id,
                        material_type=material_type,
                        title=reading_material.title,
                        order=CourseMaterial.objects.count() + 1  # increment order automatically
                    )
                    reading_material.material = course_material
                    reading_material.save()

        # Handle manual reading materials
        reading_material_titles = request.POST.getlist('reading_material_title[]')
        reading_material_contents = request.POST.getlist('reading_material_content[]')
        reading_material_types = request.POST.getlist('reading_material_type[]')
        for title, content, material_type in zip(reading_material_titles, reading_material_contents, reading_material_types):
            if title and content and material_type:
                reading_material = ReadingMaterial.objects.create(
                    title=title,
                    content=content
                )
                course_material = CourseMaterial.objects.create(
                    session=session,
                    material_id=reading_material.id,
                    material_type=material_type,
                    title=reading_material.title,
                    order=CourseMaterial.objects.count() + 1  # increment order automatically
                )
                reading_material.material = course_material
                reading_material.save()

        messages.success(request, 'Course content updated successfully.')
        return redirect(reverse('course:course_content_edit', args=[course.pk, session.id]))

    # Context to render the template
    context = {
        'course': course,
        'sessions': sessions,
        'selected_session': session,
        'reading_materials': reading_materials,
        'material_types': dict(CourseMaterial.MATERIAL_TYPE_CHOICES),
    }

    return render(request, 'material/course_content_edit.html', context)

@login_required
def toggle_publish(request, pk):
    course = get_object_or_404(Course, pk=pk)
    if request.user == course.instructor or request.user.is_superuser:
        course.published = not course.published
        course.save()
    return redirect('course:course_detail', pk=pk)

@login_required
def generate_certificate_png(request, pk):
    course = get_object_or_404(Course, pk=pk)
    student = request.user

    # Verify that the student has completed the course
    sessions = Session.objects.filter(course=course).count()
    completed_sessions = SessionCompletion.objects.filter(
        course=course,
        user=student,
        completed=True
    ).distinct().count()

    if completed_sessions != sessions:
        return HttpResponse("You have not completed this course yet.", status=403)

    # Dynamically find the background image in the course app's static directory
    app_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))  # Root of the project
    background_image_path = os.path.join(app_dir, 'course', 'static', 'course', 'images', 'certificate_background.jpg')

    if os.path.exists(background_image_path):
        with open(background_image_path, "rb") as image_file:
            encoded_string = base64.b64encode(image_file.read()).decode()
    else:
        return HttpResponse(f"Background image not found at {background_image_path}", status=500)

    # Generate the certificate
    context = {
        'student_name': student.get_full_name() or student.username,
        'course_name': course.course_name,
        'completion_date': datetime.now().strftime("%B %d, %Y"),
        'background_image_base64': encoded_string,
    }

    return render(request, 'course/certificate_template.html', context)

# Views for Topics
def topic_list(request):
    module_groups = ModuleGroup.objects.all()
    topics = Topic.objects.all()
    return render(request, 'topic-tag/topic_list.html', {'module_groups': module_groups, 'topics': topics})

def topic_add(request):
    if request.method == 'POST':
        form = TopicForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Topic added successfully.')
            return redirect('course:topic_list')
    else:
        form = TopicForm()
    return render(request, 'topic-tag/topic_form.html', {'form': form, 'title': 'Add Topic'})

def topic_edit(request, pk):
    topic = get_object_or_404(Topic, pk=pk)
    if request.method == 'POST':
        form = TopicForm(request.POST, instance=topic)
        if form.is_valid():
            form.save()
            messages.success(request, 'Topic updated successfully.')
            return redirect('course:topic_list')
    else:
        form = TopicForm(instance=topic)
    return render(request, 'topic-tag/topic_form.html', {'form': form, 'title': 'Edit Topic'})

def topic_delete(request, pk):
    topic = get_object_or_404(Topic, pk=pk)
    if request.method == 'POST':
        topic.delete()
        messages.success(request, 'Topic deleted successfully.')
        return redirect('course:topic_list')
    return render(request, 'topic-tag/topic_confirm_delete.html', {'object': topic, 'title': 'Delete Topic'})


# Views for Tags
def tag_list(request):
    module_groups = ModuleGroup.objects.all()
    tags = Tag.objects.all()
    return render(request, 'topic-tag/tag_list.html', {'tags': tags, 'module_groups': module_groups,})

def tag_add(request):
    if request.method == 'POST':
        form = TagForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tag added successfully.')
            return redirect('course:tag_list')
    else:
        form = TagForm()
    return render(request, 'topic-tag/tag_form.html', {'form': form, 'title': 'Add Tag'})

def tag_edit(request, pk):
    tag = get_object_or_404(Tag, pk=pk)
    if request.method == 'POST':
        form = TagForm(request.POST, instance=tag)
        if form.is_valid():
            form.save()
            messages.success(request, 'Tag updated successfully.')
            return redirect('course:tag_list')
    else:
        form = TagForm(instance=tag)
    return render(request, 'topic-tag/tag_form.html', {'form': form, 'title': 'Edit Tag'})

def tag_delete(request, pk):
    tag = get_object_or_404(Tag, pk=pk)
    if request.method == 'POST':
        tag.delete()
        messages.success(request, 'Tag deleted successfully.')
        return redirect('course:tag_list')
    return render(request, 'topic-tag/tag_confirm_delete.html', {'object': tag, 'title': 'Delete Tag'})