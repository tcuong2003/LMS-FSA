from django.shortcuts import render, get_object_or_404, redirect
from .models import Quiz, Question, AnswerOption, StudentQuizAttempt, StudentAnswer
from .forms import QuizForm, QuestionForm, AnswerOptionForm, QuizAnswerForm,ExcelUploadForm
from module_group.models import ModuleGroup
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from course.models import Course
from subject.models import Subject
import json
import csv
import os
from django.conf import settings
import openpyxl
from django.utils import timezone
from openpyxl.utils import get_column_letter



def quiz_list(request):
    module_groups = ModuleGroup.objects.all()
    quizzes = Quiz.objects.select_related('course').all().order_by('-created_at')
    courses = Course.objects.all()
    # Lọc quiz dựa trên subject được chọn
    selected_course = request.GET.get('course', '')
    if selected_course:
        quizzes = quizzes.filter(course__id=selected_course)

    context = {
        'module_groups': module_groups,
        'quizzes': quizzes,
        'courses': courses,
        'selected_course': selected_course,
    }
    return render(request, 'quiz_list.html', context)

def quiz_add(request):
    if request.method == 'POST':
        form = QuizForm(request.POST)
        if form.is_valid():
            quiz = form.save(commit=False)  # Không lưu ngay lập tức
            quiz.start_datetime = request.POST.get('start_datetime')  
            quiz.end_datetime = request.POST.get('end_datetime')  
            quiz.attempts_allowed = request.POST.get('attempts_allowed')  
            quiz.save()  
            return redirect('quiz:quiz_list')
        else:
            print(form.errors)  # Có thể dùng logging thay vì print
    else:
        form = QuizForm()
    return render(request, 'quiz_form.html', {'form': form})


def quiz_edit(request, pk):
    quiz = get_object_or_404(Quiz, pk=pk)
    if request.method == 'POST':
        form = QuizForm(request.POST, instance=quiz)
        if form.is_valid():
            quiz = form.save(commit=False)
            quiz.start_datetime = request.POST.get('start_datetime')  
            quiz.end_datetime = request.POST.get('end_datetime')  
            quiz.attempts_allowed = request.POST.get('attempts_allowed')  
            quiz.save()  
            return redirect('quiz:quiz_list')
    else:
        form = QuizForm(instance=quiz)
    return render(request, 'quiz_form.html', {'form': form})

def quiz_delete(request, pk):
    quiz = get_object_or_404(Quiz, pk=pk)
    if request.method == 'POST':
        quiz.delete()
        return redirect('quiz:quiz_list')
    return render(request, 'quiz_confirm_delete.html', {'quiz': quiz})

def quiz_question(request, pk):
    quiz = get_object_or_404(Quiz, pk=pk)
    module_groups = ModuleGroup.objects.all()
    questions = Question.objects.filter(quiz=quiz)
    return render(request, 'quiz_question.html', {'quiz': quiz, 'questions': questions, 'module_groups': module_groups})

def quiz_detail(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    questions = quiz.questions.all()
    return render(request, 'quiz_detail.html', {'quiz': quiz, 'questions': questions})

def question_add(request, quiz_id):
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.quiz = quiz
            question.save()
            return redirect('quiz:quiz_question', pk=quiz.id)
    else:
        form = QuestionForm()
    return render(request, 'question_form.html', {'quiz': quiz, 'form': form})

def question_edit(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    quiz = get_object_or_404(Quiz, id=question.quiz.id)  # Lấy quiz từ question

    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            return redirect('quiz:quiz_detail', pk=quiz.id)  # Sử dụng quiz.id
    else:
        form = QuestionForm(instance=question)

    return render(request, 'question_form.html', {'quiz': quiz, 'form': form})

def question_delete(request, pk):
    question = get_object_or_404(Question, pk=pk)
    quiz_id = question.quiz.id
    if request.method == 'POST':
        question.delete()   
        return redirect('quiz:quiz_detail', pk=question.quiz.id)
    return render(request, 'question_confirm_delete.html', {'question': question})



def question_detail(request, pk):
    question = get_object_or_404(Question, pk=pk)
    answer_options = AnswerOption.objects.filter(question=question)
    context = {
        'question': question,
        'answer_options': answer_options
    }
    return render(request, 'question_detail.html', context)


def answer_option_add(request, question_pk):
    question = get_object_or_404(Question, pk=question_pk)

    if request.method == 'POST':
        option_count = int(request.POST.get('option_count', 0))
        
        # Xử lý các tùy chọn đã bị xóa
        removed_options = request.POST.get('removed_options', '')
        if removed_options:
            removed_option_ids = removed_options.split(',')
            AnswerOption.objects.filter(id__in=removed_option_ids).delete()

        # Lấy danh sách option_text hiện tại để kiểm tra trùng lặp
        existing_options = set(question.answer_options.values_list('option_text', flat=True))
        
        for i in range(option_count):
            option_text = request.POST.get(f'option_text_{i}')
            is_correct = request.POST.get(f'is_correct_{i}', False) == 'on'
            if option_text and option_text not in existing_options:  # Kiểm tra trùng lặp
                AnswerOption.objects.create(question=question, option_text=option_text, is_correct=is_correct)

        return redirect('quiz:question_detail', pk=question.pk)

    # Lấy các answer options đã tồn tại
    answer_options = question.answer_options.all()

    return render(request, 'answer_option_form.html', {
        'question': question,
        'options_range': range(4),
        'answer_options': answer_options  # Truyền answer options vào template
    })


def answer_option_edit(request, pk):
    option = get_object_or_404(AnswerOption, pk=pk)
    if request.method == 'POST':
        form = AnswerOptionForm(request.POST, instance=option)
        if form.is_valid():
            form.save()
            return redirect('quiz:question_detail', pk=option.question.id)
    else:
        form = AnswerOptionForm(instance=option)
    return render(request, 'answer_option_form.html', {'form': form})

def answer_option_delete(request, pk):
    option = get_object_or_404(AnswerOption, pk=pk)
    if request.method == 'POST':
        question_id = option.question.id
        option.delete()
        return redirect('quiz:question_detail', pk=question_id)
    return render(request, 'answer_option_form.html', {'option': option})


@login_required
def take_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    questions = quiz.questions.all()  # Lấy tất cả các câu hỏi trong quiz

    if request.method == 'POST':
        # Khi người dùng gửi form (submit quiz)
        with transaction.atomic():
            # Tạo một lần thử quiz mới cho học sinh
            attempt = StudentQuizAttempt.objects.create(user=request.user, quiz=quiz, score=0.0)

            total_score = 0
            for question in questions:
                selected_option_id = request.POST.get(f'question_{question.id}')  # Lấy lựa chọn được chọn cho câu hỏi
                text_response = request.POST.get(f'text_response_{question.id}')  # Thay đổi tên ở đây
                # Nếu câu hỏi là một câu hỏi dạng text, không cần truy xuất selected_option
                selected_option = None
                if selected_option_id and selected_option_id.isdigit():
                    selected_option = AnswerOption.objects.get(id=int(selected_option_id))
                # Lưu câu trả lời của học sinh
                StudentAnswer.objects.create(
                    attempt=attempt,
                    question=question,
                    selected_option=selected_option,
                    text_response=text_response 
                )

                # Kiểm tra nếu lựa chọn đúng, cộng điểm
                if selected_option and selected_option.is_correct:
                    total_score += question.points

            # Cập nhật điểm tổng của lần thử
            attempt.score = total_score
            attempt.save()

            return redirect('quiz:quiz_result', quiz_id=quiz.id, attempt_id=attempt.id)  # Chuyển tới trang kết quả sau khi nộp bài

    # Render trang quiz với câu hỏi và lựa chọn
    return render(request, 'take_quiz.html', {'quiz': quiz, 'questions': questions})


@login_required
def quiz_result(request, quiz_id, attempt_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id, user=request.user)
    student_answers = StudentAnswer.objects.filter(attempt=attempt)

    return render(request, 'quiz_result.html', {
        'quiz' : quiz,
        'attempt': attempt,
        'student_answers': student_answers,
    })

def import_questions(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    if request.method == 'POST':
        excel_file = request.FILES['file']
        
        # Đọc file Excel với tất cả các cột dưới dạng chuỗi
        df = pd.read_excel(excel_file, dtype=str)  # Đảm bảo tất cả đều là chuỗi
        
        # Duyệt qua từng dòng trong DataFrame và lưu vào cơ sở dữ liệu
        for index, row in df.iterrows():
            question_text = row['Question']
            correct_answer = row['Correct Answer']
            question_type = row['Question Type']
            
            # Xác định các lựa chọn đáp án dựa trên loại câu hỏi
            if question_type == 'MCQ':
                answers = {
                    'A': row['A'],
                    'B': row['B'],
                    'C': row['C'],
                    'D': row['D'],
                }
            elif question_type == 'TF':
                answers = {
                    'A': row['A'],
                    'B': row['B']
                }
            elif question_type == 'TEXT':
                answers = {}  # Không cần đáp án cho câu hỏi văn bản
            else:
                answers = {}  # Xử lý các loại khác nếu cần

            # Tạo đối tượng Question
            question = Question.objects.create(
                quiz=quiz,  
                question_text=question_text,
                question_type=question_type,  # Đặt loại từ file Excel
                points=1  # Đặt giá trị điểm mặc định hoặc tùy chỉnh nếu cần
            )
            
            # Tạo đối tượng AnswerOption cho MCQ và TF
            if question_type in ['MCQ', 'TF']:
                for option, answer_text in answers.items():
                    answer_text = str(answer_text).strip()  # Loại bỏ khoảng trắng
                    is_correct = (option == correct_answer)  # Xác định nếu đáp án này đúng

                    # Tạo AnswerOption chỉ khi answer_text không rỗng
                    if answer_text:
                        AnswerOption.objects.create(
                            question=question, 
                            option_text=answer_text, 
                            is_correct=is_correct
                        )

        return redirect('quiz:quiz_list')  # Chuyển hướng về danh sách quiz sau khi nhập thành công

    return render(request, 'import_questions.html', {'quiz':quiz})




def export_questions(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    questions = Question.objects.filter(quiz=quiz).distinct()  # Sử dụng distinct để tránh lặp câu hỏi

    # Tạo một workbook mới
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = quiz.quiz_title

    # Viết tiêu đề cột
    sheet.append(['Question', 'Answer A', 'Answer B', 'Answer C', 'Answer D', 'Correct Answer', 'Question Type', ])

    for question in questions:
        # Lấy danh sách các đáp án
        answers = AnswerOption.objects.filter(question=question)  
        correct_answer = answers.filter(is_correct=True).first()  
        answer_list = [answer.option_text for answer in answers]

        
        if question.question_type == 'MCQ' and len(answer_list) >= 4:
            sheet.append([
                question.question_text,
                answer_list[0],  
                answer_list[1],  
                answer_list[2],  
                answer_list[3],  
                correct_answer.option_text if correct_answer else '',  
                question.question_type 
            ])
        elif question.question_type == 'TF':  
            sheet.append([
                question.question_text,
                'TRUE',  
                'FALSE',  
                '',  
                '',  
                correct_answer.option_text if correct_answer else '', 
                question.question_type  
            ])
        elif question.question_type == 'TEXT':  
            sheet.append([
                question.question_text,
                '', 
                '',  
                '',  
                '', 
                '',  
                question.question_type 
            ])

    # Tạo phản hồi HTTP với tệp Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{quiz.quiz_title}_questions.xlsx"'
    workbook.save(response)

    return response



def export_quizzes(request, format):
    quizzes = Quiz.objects.all()
    
    if format == 'json':
        # Export quizzes as JSON
        quizzes_data = []
        for quiz in quizzes:
            quizzes_data.append({
                "title": quiz.quiz_title,
                "description": quiz.quiz_description,
                "course": quiz.course.course_name if quiz.course else None,
                "total_marks": quiz.total_marks,
                "time_limit": quiz.time_limit,
                'start_datetime': quiz.start_datetime.isoformat() if quiz.start_datetime else None,
                'end_datetime': quiz.end_datetime.isoformat() if quiz.end_datetime else None,
                "attempts_allowed": quiz.attempts_allowed,
                "created_by": quiz.created_by.username if quiz.created_by else None,
                'created_at': quiz.created_at.isoformat(),  # Chuyển đổi sang chuỗi
                'updated_at': quiz.updated_at.isoformat(),  # Chuyển đổi sang chuỗi
            })
        
        response = HttpResponse(json.dumps(quizzes_data, indent=4), content_type="application/json")
        response['Content-Disposition'] = f'attachment; filename="quizzes_{timezone.now().strftime("%Y%m%d")}.json"'
        return response
    
    elif format == 'excel':
        # Export quizzes as Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="quizzes_{timezone.now().strftime("%Y%m%d")}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quizzes"

        # Write the header row
        headers = ['Title', 'Description', 'Course', 'Total Marks', 'Time Limit', 'Start Date', 'End Date', 'Attempts Allowed', 'Created By']
        ws.append(headers)

        # Write the data rows
        for quiz in quizzes:
            ws.append([
                quiz.quiz_title,
                quiz.quiz_description,
                quiz.course.course_name if quiz.course else '',
                quiz.total_marks,
                quiz.time_limit,
                quiz.start_datetime.strftime('%Y-%m-%d %H:%M') if quiz.start_datetime else '',
                quiz.end_datetime.strftime('%Y-%m-%d %H:%M') if quiz.end_datetime else '',
                quiz.attempts_allowed,
                quiz.created_by.username if quiz.created_by else ''
            ])

        # Set the column width to auto-fit content
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error for non-string content
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook to the HttpResponse
        wb.save(response)

        return response
    
    elif format == 'csv':
        # Export quizzes as CSV (Excel-readable format)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="quizzes_{timezone.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Title', 'Description', 'Course', 'Total Marks', 'Time Limit', 'Start Date', 'End Date', 'Attempts Allowed', 'Created By'])

        for quiz in quizzes:
            writer.writerow([
                quiz.quiz_title,
                quiz.quiz_description,
                quiz.course.course_name if quiz.course else '',
                quiz.total_marks,
                quiz.time_limit,
                quiz.start_datetime,
                quiz.end_datetime,
                quiz.attempts_allowed,
                quiz.created_by.username if quiz.created_by else ''
            ])

        return response

    else:
        return HttpResponse(status=400)  # Bad request if format is not supported
    


def import_quizzes(request):
    if request.method == 'POST':
        # Kiểm tra xem người dùng có tải file lên hay không
        if request.FILES.get('file'):
            file = request.FILES['file']
            if file.name.endswith('.json'):
                # Nhập từ file JSON
                data = json.load(file)
                for item in data:
                    course, created = Course.objects.get_or_create(course_name=item['course'])
                    Quiz.objects.create(
                        course=course,
                        quiz_title=item['title'],
                        quiz_description=item['description'],
                        total_marks=item['total_marks'],
                        time_limit=item['time_limit'],
                        start_datetime=item.get('start_datetime'),
                        end_datetime=item.get('end_datetime'),
                        attempts_allowed=item['attempts_allowed'],
                        created_by=request.user,
                    )
            elif file.name.endswith('.csv'):
                reader = csv.DictReader(file.read().decode('utf-8').splitlines())
                for row in reader:
                    course, created = Course.objects.get_or_create(course_name=row['Course'])
                    Quiz.objects.create(
                        course=course,
                        quiz_title=row['Title'],
                        quiz_description=row['Description'],
                        total_marks=row['Total Marks'],
                        time_limit=row['Time Limit'],
                        start_datetime=row['Start Date'],
                        end_datetime=row['End Date'],
                        attempts_allowed=row['Attempts Allowed'],
                        created_by=request.user,
                    )
            
            elif file.name.endswith('.xls') or file.name.endswith('.xlsx'):
                # Đọc dữ liệu từ file Excel sử dụng pandas
                df = pd.read_excel(file)

                # In nội dung của file để kiểm tra
                print(df.head())

                # Lặp qua từng hàng trong DataFrame
                for index, row in df.iterrows():
                    course, created = Course.objects.get_or_create(course_name=row['Course'])
                    Quiz.objects.create(
                        course=course,
                        quiz_title=row['Title'],
                        quiz_description=row['Description'],
                        total_marks=row['Total Marks'],
                        time_limit=row['Time Limit'],
                        start_datetime=row['Start Date'],
                        end_datetime=row['End Date'],
                        attempts_allowed=row['Attempts Allowed'],
                        created_by=request.user,
                    )
    
            # Thêm việc xử lý file từ static nếu cần
            elif 'static_file' in request.POST:
                file_path = os.path.join(settings.STATIC_ROOT, 'excel', 'import_quiz.csv')
                with open(file_path, newline='', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        course, created = Course.objects.get_or_create(course_name=row['Course'])
                        Quiz.objects.create(
                            course=course,
                            quiz_title=row['Title'],
                            quiz_description=row['Description'],
                            total_marks=row['Total Marks'],
                            time_limit=row['Time Limit'],
                            start_datetime=row['Start Date'],
                            end_datetime=row['End Date'],
                            attempts_allowed=row['Attempts Allowed'],
                            created_by=request.user,
                        )
            return redirect('quiz:quiz_list')

        return render(request, 'import_quizzes.html')
    
#-----------Convert-------------
import zipfile
def excel_to_json_view(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            json_files = []

            for excel_file in form.cleaned_data['files']:  # Sử dụng cleaned_data
                try:
                    # Đọc dữ liệu từ file Excel
                    excel_data = pd.read_excel(excel_file, sheet_name=None)
                    print(f"EXCEL_DATA IS :{excel_data}")
                    json_data = {}

                    for sheet_name, df in excel_data.items():
                        # Chuyển đổi dữ liệu của mỗi sheet thành JSON
                        json_data[sheet_name] = df.to_dict(orient='records')
                    print(f"json_data is {json_data}")
                    # Tạo tệp JSON cho từng tệp Excel
                    json_filename = f"{excel_file.name.split('.')[0]}.json"
                    json_string = json.dumps(json_data, indent=4, ensure_ascii=False)
                    print(f"JSON_STRING IS {json_string}")
                    json_files.append((json_filename, json_string))
                    print(f"son_files is {json_files}")

                except Exception as e:
                    print(f"Lỗi khi xử lý tệp Excel '{excel_file.name}': {e}")

            # Tạo tệp ZIP
            zip_filename = 'converted_data.zip'
            response = HttpResponse(content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'

            with zipfile.ZipFile(response, 'w') as zip_file:
                for json_filename, json_string in json_files:
                    # Thêm tệp JSON vào tệp ZIP
                    zip_file.writestr(json_filename, json_string)

            return response

    else:
        form = ExcelUploadForm()

    return render(request, 'excel_to_json.html', {'form': form})