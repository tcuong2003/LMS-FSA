from django.shortcuts import render, get_object_or_404, redirect
from .models import Quiz, Question, AnswerOption, StudentQuizAttempt, StudentAnswer
from .forms import QuizForm, QuestionForm, AnswerOptionForm, QuizAnswerForm, ExcelUploadForm
from module_group.models import ModuleGroup
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse
import pandas as pd
from django.http import HttpResponse
from openpyxl import Workbook
from course.models import Course
from subject.models import Subject
import json
import csv
import os
from django.conf import settings
import openpyxl
from django.utils import timezone
from openpyxl.utils import get_column_letter
from django.db.models import Count
from django.urls import reverse
from django.core.mail import send_mail
from django.contrib.auth.decorators import login_required
from django.contrib.auth import authenticate, login
from user.models import User



def quiz_list(request):
    module_groups = ModuleGroup.objects.all()
    quizzes = Quiz.objects.select_related('course').annotate(question_count=Count('questions')).all().order_by('-created_at')
    
    courses = Course.objects.all()
    # Lọc quiz dựa trên subject được chọn
    selected_course = request.GET.get('course', '')
    if selected_course:
        quizzes = quizzes.filter(course__id=selected_course)

    invited_count = 50
    assessed_count = 50
    qualified_count = 30  # Hoặc lấy từ cơ sở dữ liệu nếu cần
    qualifying_percentage = (qualified_count / invited_count) * 100 if invited_count > 0 else 0
    qualifying_percentage = f"{int(qualifying_percentage)}"
    
    context = {
        'module_groups': module_groups,
        'quizzes': quizzes,
        'courses': courses,
        'selected_course': selected_course,
        'invited_count': invited_count,
        'assessed_count': assessed_count,
        'qualified_count': qualified_count,
        'qualifying_percentage': qualifying_percentage,
    }
    return render(request, 'quiz_list.html', context)

def quiz_add(request):
    if request.method == 'POST':
        form = QuizForm(request.POST)
        if form.is_valid():
            quiz = form.save(commit=False)  # Không lưu ngay lập tức
            quiz.start_datetime = request.POST.get('start_datetime')  
            quiz.end_datetime = request.POST.get('end_datetime')  
            quiz.attempts_allowed = request.POST.get('attempts_allowed')  
            quiz.created_by = request.user
            quiz.save()  
            return redirect('quiz:quiz_list')
        else:
            print(form.errors)  # Có thể dùng logging thay vì print
    else:
        form = QuizForm()
    return render(request, 'quiz_form.html', {'form': form})


def quiz_edit(request, pk):
    quiz = get_object_or_404(Quiz, pk=pk)
    if request.method == 'POST':
        form = QuizForm(request.POST, instance=quiz)
        if form.is_valid():
            quiz = form.save(commit=False)
            quiz.start_datetime = request.POST.get('start_datetime')  
            quiz.end_datetime = request.POST.get('end_datetime')  
            quiz.attempts_allowed = request.POST.get('attempts_allowed')  
            quiz.save()  
            return redirect('quiz:quiz_list')
    else:
        form = QuizForm(instance=quiz)
    return render(request, 'quiz_form.html', {'form': form, 'quiz': quiz})

def quiz_delete(request, pk):
    quiz = get_object_or_404(Quiz, pk=pk)
    if request.method == 'POST':
        quiz.delete()
        return redirect('quiz:quiz_list')
    return render(request, 'quiz_confirm_delete.html', {'quiz': quiz})
    

def quiz_question(request, pk):
    quiz = get_object_or_404(Quiz, pk=pk)
    
    search_query = request.GET.get('search', '')

    if search_query:  
        questions = Question.objects.filter(quiz=quiz, question_text__icontains=search_query).prefetch_related('answer_options')
    else:
        questions = Question.objects.filter(quiz=quiz).prefetch_related('answer_options')


    question_form = QuestionForm()

    if request.method == 'POST':
        if 'add_question' in request.POST:  # Check if submitting a new question
            form = QuestionForm(request.POST)
            if form.is_valid():
                question = form.save(commit=False)
                question.quiz = quiz
                question.save()
                return redirect('quiz:quiz_question', pk=quiz.id)  # Redirect to the same page to refresh questions
            
        elif 'edit_question' in request.POST:  # Handle question editing
            question_id = request.POST.get('question_id')
            question = get_object_or_404(Question, pk=question_id)
            question.question_text = request.POST.get('question_text')
            question.question_type = request.POST.get('question_type')
            question.points = request.POST.get('points')
            question.save()
            return redirect('quiz:quiz_question', pk=quiz.id) # Redirect after edit    

        elif 'edit_answers' in request.POST:
            question_id = request.POST.get('question_id')
            question = get_object_or_404(Question, pk=question_id)

            # Handle existing answers (edits)
            answer_ids = request.POST.getlist('answer_id[]')
            option_texts = request.POST.getlist('option_text[]')
            is_correct_list = request.POST.getlist('is_correct[]')

            for i in range(len(option_texts)):
                answer_id = answer_ids[i]  # answer_id is relevant here
                is_correct = answer_id in is_correct_list
                
                if answer_id:
                    try:
                        answer = AnswerOption.objects.get(id=answer_id)
                        answer.option_text = option_texts[i]
                        answer.is_correct = is_correct
                        answer.save()
                    except AnswerOption.DoesNotExist:
                        pass  # Handle if answer was deleted

                # Handle new answers (separate loop)
            new_option_texts = request.POST.getlist('new_option_text[]')
            new_is_correct_list = request.POST.getlist('new_is_correct[]')

            for i, option_text in enumerate(new_option_texts):
                is_correct = 'new_' + str(i) in new_is_correct_list  # Check if the index (as string) is in new_is_correct_list
                AnswerOption.objects.create(question=question, option_text=option_text, is_correct=is_correct)
                            

            # Handle answer removal
            removed_answer_ids_str = request.POST.get('removed_answers')
            if removed_answer_ids_str:
                removed_answer_ids = removed_answer_ids_str.split(',')  # Split the string
                for answer_id_str in removed_answer_ids:
                    try:
                        answer_id = int(answer_id_str)
                        AnswerOption.objects.get(id=answer_id).delete()
                    except (ValueError, AnswerOption.DoesNotExist):
                        pass  # Handle errors (e.g., invalid ID, already deleted)

            return redirect('quiz:quiz_question', pk=quiz.id)
        
    return render(request, 'quiz_question.html', {
        'quiz': quiz,
        'questions': questions,
        'question_form': question_form,
        'search_query': search_query,
    })




def get_answers(request, question_pk):
    question = get_object_or_404(Question, pk=question_pk)
    answers = question.answer_options.all()
    return JsonResponse({'answers': list(answers.values())}) # Serialize to JSON

def quiz_detail(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    questions = quiz.questions.all()
    return render(request, 'quiz_detail.html', {'quiz': quiz, 'questions': questions})

def question_add(request, quiz_id):
    quiz = get_object_or_404(Quiz, pk=quiz_id)
    if request.method == 'POST':
        form = QuestionForm(request.POST)
        if form.is_valid():
            question = form.save(commit=False)
            question.quiz = quiz
            question.save()
            return redirect('quiz:quiz_question', pk=quiz.id)
    else:
        form = QuestionForm()
    return render(request, 'question_form.html', {'quiz': quiz, 'form': form})

def question_edit(request, question_id): # Now takes question_id directly
    question = get_object_or_404(Question, id=question_id)
    quiz = question.quiz
    # Set the editing flag in session so that the template shows the form
    request.session['editing_question'] = question.id

    if request.method == 'POST':
        form = QuestionForm(request.POST, instance=question)
        if form.is_valid():
            form.save()
            del request.session['editing_question'] # Clear the edit flag from session
            return redirect('quiz:quiz_question', pk=quiz.id)  
    else:
        form = QuestionForm(instance=question)
    
    return redirect(reverse('quiz:quiz_question', kwargs={'pk': quiz.id}))

def question_delete(request, pk):
    question = get_object_or_404(Question, pk=pk)
    quiz_id = question.quiz.id
    if request.method == 'POST':
        question.delete()   
        return redirect(reverse('quiz:quiz_question', kwargs={'pk': quiz_id})) # Corrected redirect
    return render(request, 'question_confirm_delete.html', {'question': question})


def question_detail(request, pk):
    question = get_object_or_404(Question, pk=pk)
    answer_options = AnswerOption.objects.filter(question=question)
    context = {
        'question': question,
        'answer_options': answer_options
    }
    return render(request, 'question_detail.html', context)



def answer_option_edit(request, pk):
    option = get_object_or_404(AnswerOption, pk=pk)
    if request.method == 'POST':
        form = AnswerOptionForm(request.POST, instance=option)
        if form.is_valid():
            form.save()
            return redirect('quiz:question_detail', pk=option.question.id)
    else:
        form = AnswerOptionForm(instance=option)
    return render(request, 'answer_option_form.html', {'form': form})

def answer_option_delete(request, pk):
    option = get_object_or_404(AnswerOption, pk=pk)
    if request.method == 'POST':
        question_id = option.question.id
        option.delete()
        return redirect('quiz:question_detail', pk=question_id)
    return render(request, 'answer_option_form.html', {'option': option})


@login_required
def take_quiz(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    questions = quiz.questions.all()  # Lấy tất cả các câu hỏi trong quiz
    total_marks = quiz.total_marks
    total_questions = questions.count()

    if request.method == 'POST':
        # Khi người dùng gửi form (submit quiz)
        with transaction.atomic():
            # Tạo một lần thử quiz mới cho học sinh
            attempt = StudentQuizAttempt.objects.create(user=request.user, quiz=quiz, score=0.0)

            correct_answers = 0
            for question in questions:
                selected_option_id = request.POST.get(f'question_{question.id}')  # Lấy lựa chọn được chọn cho câu hỏi
                text_response = request.POST.get(f'text_response_{question.id}')  # Thay đổi tên ở đây
                # Nếu câu hỏi là một câu hỏi dạng text, không cần truy xuất selected_option
                selected_option = None
                if selected_option_id and selected_option_id.isdigit():
                    selected_option = AnswerOption.objects.get(id=int(selected_option_id))
                # Lưu câu trả lời của học sinh
                StudentAnswer.objects.create(
                    attempt=attempt,
                    question=question,
                    selected_option=selected_option,
                    text_response=text_response 
                )

                # Kiểm tra nếu lựa chọn đúng, cộng điểm
                if selected_option and selected_option.is_correct:
                    correct_answers += 1

            final_score = (total_marks / total_questions) * correct_answers

            # Cập nhật điểm cho lần thử
            attempt.score = final_score
            attempt.save()

            return redirect('quiz:quiz_result', quiz_id=quiz.id, attempt_id=attempt.id)  # Chuyển tới trang kết quả sau khi nộp bài

    # Render trang quiz với câu hỏi và lựa chọn
    return render(request, 'take_quiz.html', {'quiz': quiz, 'questions': questions})


@login_required
def quiz_result(request, quiz_id, attempt_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    attempt = get_object_or_404(StudentQuizAttempt, id=attempt_id, user=request.user)
    context = _get_quiz_result_context(quiz, attempt) # Sử dụng hàm chung
    return render(request, 'quiz_result.html', context)

def import_questions(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    if request.method == 'POST':
        excel_file = request.FILES['file']
        
        # Đọc file Excel với tất cả các cột dưới dạng chuỗi
        df = pd.read_excel(excel_file, dtype=str)  # Đảm bảo tất cả đều là chuỗi
        
        # Duyệt qua từng dòng trong DataFrame và lưu vào cơ sở dữ liệu
        for index, row in df.iterrows():
            question_text = row['Question']
            correct_answer = row['Correct Answer']
            question_type = row['Question Type']
            
            # Xác định các lựa chọn đáp án dựa trên loại câu hỏi
            if question_type == 'MCQ':
                answers = {
                    'A': row['A'],
                    'B': row['B'],
                    'C': row['C'],
                    'D': row['D'],
                }
            elif question_type == 'TF':
                answers = {
                    'A': row['A'],
                    'B': row['B']
                }
            elif question_type == 'TEXT':
                answers = {}  # Không cần đáp án cho câu hỏi văn bản
            else:
                answers = {}  # Xử lý các loại khác nếu cần

            # Tạo đối tượng Question
            question = Question.objects.create(
                quiz=quiz,  
                question_text=question_text,
                question_type=question_type,  # Đặt loại từ file Excel
                points=1  # Đặt giá trị điểm mặc định hoặc tùy chỉnh nếu cần
            )
            
            # Tạo đối tượng AnswerOption cho MCQ và TF
            if question_type in ['MCQ', 'TF']:
                for option, answer_text in answers.items():
                    answer_text = str(answer_text).strip()  # Loại bỏ khoảng trắng
                    is_correct = (option == correct_answer)  # Xác định nếu đáp án này đúng

                    # Tạo AnswerOption chỉ khi answer_text không rỗng
                    if answer_text:
                        AnswerOption.objects.create(
                            question=question, 
                            option_text=answer_text, 
                            is_correct=is_correct
                        )

        return redirect('quiz:quiz_list')  # Chuyển hướng về danh sách quiz sau khi nhập thành công

    return render(request, 'import_questions.html', {'quiz':quiz})




def export_questions(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    questions = Question.objects.filter(quiz=quiz).distinct()  # Sử dụng distinct để tránh lặp câu hỏi

    # Tạo một workbook mới
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = quiz.quiz_title

    # Viết tiêu đề cột
    sheet.append(['Question', 'Answer A', 'Answer B', 'Answer C', 'Answer D', 'Correct Answer', 'Question Type', ])

    for question in questions:
        # Lấy danh sách các đáp án
        answers = AnswerOption.objects.filter(question=question)  
        correct_answer = answers.filter(is_correct=True).first()  
        answer_list = [answer.option_text for answer in answers]

        
        if question.question_type == 'MCQ' and len(answer_list) >= 4:
            sheet.append([
                question.question_text,
                answer_list[0],  
                answer_list[1],  
                answer_list[2],  
                answer_list[3],  
                correct_answer.option_text if correct_answer else '',  
                question.question_type 
            ])
        elif question.question_type == 'TF':  
            sheet.append([
                question.question_text,
                'TRUE',  
                'FALSE',  
                '',  
                '',  
                correct_answer.option_text if correct_answer else '', 
                question.question_type  
            ])
        elif question.question_type == 'TEXT':  
            sheet.append([
                question.question_text,
                '', 
                '',  
                '',  
                '', 
                '',  
                question.question_type 
            ])

    # Tạo phản hồi HTTP với tệp Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{quiz.quiz_title}_questions.xlsx"'
    workbook.save(response)

    return response



def export_quizzes(request, format):
    quizzes = Quiz.objects.all()
    
    if format == 'json':
        # Export quizzes as JSON
        quizzes_data = []
        for quiz in quizzes:
            quizzes_data.append({
                "title": quiz.quiz_title,
                "description": quiz.quiz_description,
                "course": quiz.course.course_name if quiz.course else None,
                "total_marks": quiz.total_marks,
                "time_limit": quiz.time_limit,
                'start_datetime': quiz.start_datetime.isoformat() if quiz.start_datetime else None,
                'end_datetime': quiz.end_datetime.isoformat() if quiz.end_datetime else None,
                "attempts_allowed": quiz.attempts_allowed,
                "created_by": quiz.created_by.username if quiz.created_by else None,
                'created_at': quiz.created_at.isoformat(),  # Chuyển đổi sang chuỗi
                'updated_at': quiz.updated_at.isoformat(),  # Chuyển đổi sang chuỗi
            })
        
        response = HttpResponse(json.dumps(quizzes_data, indent=4), content_type="application/json")
        response['Content-Disposition'] = f'attachment; filename="quizzes_{timezone.now().strftime("%Y%m%d")}.json"'
        return response
    
    elif format == 'excel':
        # Export quizzes as Excel
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="quizzes_{timezone.now().strftime("%Y%m%d")}.xlsx"'

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Quizzes"

        # Write the header row
        headers = ['Title', 'Description', 'Course', 'Total Marks', 'Time Limit', 'Start Date', 'End Date', 'Attempts Allowed', 'Created By']
        ws.append(headers)

        # Write the data rows
        for quiz in quizzes:
            ws.append([
                quiz.quiz_title,
                quiz.quiz_description,
                quiz.course.course_name if quiz.course else '',
                quiz.total_marks,
                quiz.time_limit,
                quiz.start_datetime.strftime('%Y-%m-%d %H:%M') if quiz.start_datetime else '',
                quiz.end_datetime.strftime('%Y-%m-%d %H:%M') if quiz.end_datetime else '',
                quiz.attempts_allowed,
                quiz.created_by.username if quiz.created_by else ''
            ])

        # Set the column width to auto-fit content
        for col in ws.columns:
            max_length = 0
            column = get_column_letter(col[0].column)  # Get the column name
            for cell in col:
                try:  # Necessary to avoid error for non-string content
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

        # Save the workbook to the HttpResponse
        wb.save(response)

        return response
    
    elif format == 'csv':
        # Export quizzes as CSV (Excel-readable format)
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="quizzes_{timezone.now().strftime("%Y%m%d")}.csv"'

        writer = csv.writer(response)
        writer.writerow(['Title', 'Description', 'Course', 'Total Marks', 'Time Limit', 'Start Date', 'End Date', 'Attempts Allowed', 'Created By'])

        for quiz in quizzes:
            writer.writerow([
                quiz.quiz_title,
                quiz.quiz_description,
                quiz.course.course_name if quiz.course else '',
                quiz.total_marks,
                quiz.time_limit,
                quiz.start_datetime,
                quiz.end_datetime,
                quiz.attempts_allowed,
                quiz.created_by.username if quiz.created_by else ''
            ])

        return response

    else:
        return HttpResponse(status=400)  # Bad request if format is not supported
    


def import_quizzes(request):
    if request.method == 'POST':
        # Kiểm tra xem người dùng có tải file lên hay không
        if request.FILES.get('file'):
            file = request.FILES['file']
            if file.name.endswith('.json'):
                # Nhập từ file JSON
                data = json.load(file)
                for item in data:
                    course, created = Course.objects.get_or_create(course_name=item['course'])
                    Quiz.objects.create(
                        course=course,
                        quiz_title=item['title'],
                        quiz_description=item['description'],
                        total_marks=item['total_marks'],
                        time_limit=item['time_limit'],
                        start_datetime=item.get('start_datetime'),
                        end_datetime=item.get('end_datetime'),
                        attempts_allowed=item['attempts_allowed'],
                        created_by=request.user,
                    )
            elif file.name.endswith('.csv'):
                reader = csv.DictReader(file.read().decode('utf-8').splitlines())
                for row in reader:
                    course, created = Course.objects.get_or_create(course_name=row['Course'])
                    Quiz.objects.create(
                        course=course,
                        quiz_title=row['Title'],
                        quiz_description=row['Description'],
                        total_marks=row['Total Marks'],
                        time_limit=row['Time Limit'],
                        start_datetime=row['Start Date'],
                        end_datetime=row['End Date'],
                        attempts_allowed=row['Attempts Allowed'],
                        created_by=request.user,
                    )
            
            elif file.name.endswith('.xls') or file.name.endswith('.xlsx'):
                # Đọc dữ liệu từ file Excel sử dụng pandas
                df = pd.read_excel(file)

                # In nội dung của file để kiểm tra
                print(df.head())

                # Lặp qua từng hàng trong DataFrame
                for index, row in df.iterrows():
                    course, created = Course.objects.get_or_create(course_name=row['Course'])
                    Quiz.objects.create(
                        course=course,
                        quiz_title=row['Title'],
                        quiz_description=row['Description'],
                        total_marks=row['Total Marks'],
                        time_limit=row['Time Limit'],
                        start_datetime=row['Start Date'],
                        end_datetime=row['End Date'],
                        attempts_allowed=row['Attempts Allowed'],
                        created_by=request.user,
                    )
    
            # Thêm việc xử lý file từ static nếu cần
            elif 'static_file' in request.POST:
                file_path = os.path.join(settings.STATIC_ROOT, 'excel', 'import_quiz.csv')
                with open(file_path, newline='', encoding='utf-8') as csvfile:
                    reader = csv.DictReader(csvfile)
                    for row in reader:
                        course, created = Course.objects.get_or_create(course_name=row['Course'])
                        Quiz.objects.create(
                            course=course,
                            quiz_title=row['Title'],
                            quiz_description=row['Description'],
                            total_marks=row['Total Marks'],
                            time_limit=row['Time Limit'],
                            start_datetime=row['Start Date'],
                            end_datetime=row['End Date'],
                            attempts_allowed=row['Attempts Allowed'],
                            created_by=request.user,
                        )
            return redirect('quiz:quiz_list')

        return render(request, 'import_quizzes.html')
    

import zipfile
def excel_to_json_view(request):
    if request.method == 'POST':
        form = ExcelUploadForm(request.POST, request.FILES)
        if form.is_valid():
            json_files = []

            for excel_file in form.cleaned_data['files']:  # Sử dụng cleaned_data
                try:
                    # Đọc dữ liệu từ file Excel
                    excel_data = pd.read_excel(excel_file, sheet_name=None)
                    print(f"EXCEL_DATA IS :{excel_data}")
                    json_data = {}

                    for sheet_name, df in excel_data.items():
                        # Chuyển đổi dữ liệu của mỗi sheet thành JSON
                        json_data[sheet_name] = df.to_dict(orient='records')
                    print(f"json_data is {json_data}")
                    # Tạo tệp JSON cho từng tệp Excel
                    json_filename = f"{excel_file.name.split('.')[0]}.json"
                    json_string = json.dumps(json_data, indent=4, ensure_ascii=False)
                    json_files.append((json_filename, json_string))
                    

                except Exception as e:
                    print(f"Lỗi khi xử lý tệp Excel '{excel_file.name}': {e}")

            # Tạo tệp ZIP
            zip_filename = 'converted_data.zip'
            response = HttpResponse(content_type='application/zip')
            response['Content-Disposition'] = f'attachment; filename="{zip_filename}"'

            with zipfile.ZipFile(response, 'w') as zip_file:
                for json_filename, json_string in json_files:
                    # Thêm tệp JSON vào tệp ZIP
                    zip_file.writestr(json_filename, json_string)

            return response

    else:
        form = ExcelUploadForm()

    return render(request, 'excel_to_json.html', {'form': form})

def send_quiz_invite(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    if request.method == 'POST':
        if request.user.is_authenticated:  # Kiểm tra xem người dùng đã đăng nhập chưa
            sender_email = request.user.email # Sử dụng email của người dùng hiện tại
        
        emails = request.POST.get('emails').split(',')
        subject = f"Invitation to take quiz: {quiz.quiz_title}"
        public_link = request.build_absolute_uri(reverse('quiz:take_quiz_public', args=[quiz.id]))
        message = f"You are invited to take the quiz '{quiz.quiz_title}'. Please click the link below to access the quiz:\n{public_link}"

        for email in emails:
            try:
                send_mail(subject, message, sender_email, [email.strip()], fail_silently=False) # Gửi email
            except Exception as e:
                # Xử lý lỗi (log, hiển thị thông báo lỗi...)
                print(f"Error sending email to {email}: {e}")
                return HttpResponse("Error sending invites. Please check the email addresses and try again.")

        return HttpResponse("Invites sent successfully!")

    return HttpResponse("Invalid request.")
def copy_public_invite_link(request, quiz_id):
    quiz = get_object_or_404(Quiz, id=quiz_id)
    public_link = request.build_absolute_uri(reverse('quiz:take_quiz_public', args=[quiz.id]))
    return JsonResponse({'link': public_link})  # Trả về link dưới dạng JSON để dễ xử lý ở frontend


def take_quiz_public(request, quiz_id): # Hàm mới cho người dùng nhận link
    quiz = get_object_or_404(Quiz, id=quiz_id)
    if request.method == 'POST':
        email = request.POST.get('email') # Lấy email từ form đăng nhập
        user = User.objects.filter(email=email).first() # Tìm user theo email
        if user:
            login(request, user) # Đăng nhập user tự động
            return redirect('quiz:take_quiz_invited', quiz_id=quiz.id)  # Chuyển hướng đến hàm take_quiz hiện có
        else:
            return render(request, 'public_login.html', {'quiz_id': quiz_id, 'error': 'Invalid email'})  # Hiển thị thông báo lỗi

    return render(request, 'public_login.html', {'quiz_id': quiz_id}) # Render form đăng nhập


def _get_quiz_result_context(quiz, attempt):
    student_answers = StudentAnswer.objects.filter(attempt=attempt)
    questions_with_options = []
    for answer in student_answers:
        question = answer.question
        options = AnswerOption.objects.filter(question=question)
        questions_with_options.append({
            'question': question,
            'options': options,
            'selected_option': answer.selected_option
        })
    return {
        'quiz': quiz,
        'attempt': attempt,
        'questions_with_options': questions_with_options,
    }